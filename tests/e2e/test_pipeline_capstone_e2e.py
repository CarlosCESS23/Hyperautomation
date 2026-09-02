"""E2E do pipeline Capstone com seis bots."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

from openpyxl import (
    Workbook,
    load_workbook,
)

from src.bots.bot_classificacao_ml import (
    ConfiguracaoClassificacaoML,
    executar_bot_classificacao_ml,
)
from src.bots.bot_coleta_desktop import (
    ConfiguracaoColetaDesktop,
    executar_bot_coleta_desktop,
)
from src.bots.bot_coleta_web import (
    ConfiguracaoColetaWeb,
    executar_bot_coleta_web,
)
from src.bots.bot_consolidacao import (
    ConfiguracaoConsolidacao,
    executar_bot_consolidacao,
)
from src.bots.bot_entrada import (
    executar_bot_entrada,
)
from src.bots.bot_relatorio_capstone import (
    ConfiguracaoRelatorioCapstone,
    executar_bot_relatorio_capstone,
)
from src.contratos_capstone import (
    ArtefatoClassificacaoML,
    ArtefatoConsolidacao,
    ArtefatoEstoqueDesktop,
    ArtefatoPedidosFornecedor,
    PedidoFornecedor,
)
from src.decisao_hibrida import (
    ResultadoDecisaoHibrida,
)


EXECUTION_ID = "exec-e2e-capstone"
CORRELATION_ID = "corr-e2e-capstone"


def criar_planilha_entrada(
    caminho: Path,
) -> Path:
    """Cria uma entrada estruturalmente válida para o Bot A."""

    workbook = Workbook()

    base = workbook.active
    base.title = "Base_Referencia"

    base.cell(
        row=2,
        column=1,
        value="lote_id",
    )

    for linha, lote_id in enumerate(
        (
            "LOTE-001",
            "LOTE-002",
            "LOTE-003",
        ),
        start=3,
    ):
        base.cell(
            row=linha,
            column=1,
            value=lote_id,
        )

    inspecao = workbook.create_sheet(
        "Insp_01"
    )

    cabecalhos = (
        "lote_id",
        "produto",
        "linha",
        "turno",
        "status",
        "responsavel",
        "data",
        "observacao",
    )

    for coluna, cabecalho in enumerate(
        cabecalhos,
        start=1,
    ):
        inspecao.cell(
            row=3,
            column=coluna,
            value=cabecalho,
        )

    inspecao.append(
        (
            "LOTE-001",
            "Sensor",
            "Linha 1",
            "Manhã",
            "APROVADO",
            "Carlos",
            "01/09/2026",
            "Registro do E2E",
        )
    )

    workbook.save(caminho)
    workbook.close()

    return caminho


class AutomacaoDesktopFalsa:
    """Simula a interface visual do estoque."""

    def __init__(self) -> None:
        self._indice = 0

        self._paginas = (
            (
                "lote_id\tproduto\t"
                "quantidade_disponivel\t"
                "localizacao\tstatus_estoque\n"
                "LOTE-001\tSensor de temperatura\t"
                "10\tA-01\tDISPONIVEL\n"
                "LOTE-002\tVálvula de controle\t"
                "0\tA-02\tINDISPONIVEL"
            ),
            (
                "lote_id\tproduto\t"
                "quantidade_disponivel\t"
                "localizacao\tstatus_estoque\n"
                "LOTE-003\tCabo de instrumentação\t"
                "3\tB-01\tESTOQUE_BAIXO"
            ),
        )

    def localizar_aplicacao(
        self,
        timeout_ms: int,
    ) -> bool:
        del timeout_ms
        return True

    def copiar_pagina_visivel(
        self,
    ) -> str:
        return self._paginas[
            self._indice
        ]

    def avancar_pagina(
        self,
        timeout_ms: int,
    ) -> bool:
        del timeout_ms

        if (
            self._indice + 1
            >= len(self._paginas)
        ):
            return False

        self._indice += 1
        return True

    def capturar_screenshot(
        self,
        caminho: Path,
    ) -> None:
        caminho.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        caminho.write_bytes(b"PNG")


class AutomacaoWebFalsa:
    """Simula o portal de fornecedores."""

    def __init__(self) -> None:
        self._indice = 0

        self._paginas = (
            (
                PedidoFornecedor(
                    pedido_id="PED-001",
                    lote_id="LOTE-001",
                    fornecedor="Fornecedor A",
                    produto=(
                        "Sensor de temperatura"
                    ),
                    quantidade_pedida=6,
                    status_pedido="PENDENTE",
                ),
                PedidoFornecedor(
                    pedido_id="PED-002",
                    lote_id="LOTE-002",
                    fornecedor="Fornecedor B",
                    produto=(
                        "Válvula de controle"
                    ),
                    quantidade_pedida=2,
                    status_pedido="PENDENTE",
                ),
            ),
            (
                PedidoFornecedor(
                    pedido_id="PED-003",
                    lote_id="LOTE-003",
                    fornecedor="Fornecedor C",
                    produto=(
                        "Cabo de instrumentação"
                    ),
                    quantidade_pedida=5,
                    status_pedido="PENDENTE",
                ),
            ),
        )

    def abrir_portal(
        self,
        url: str,
        timeout_ms: int,
    ) -> None:
        del url
        del timeout_ms

    def coletar_pagina_atual(
        self,
    ):
        return self._paginas[
            self._indice
        ]

    def avancar_pagina(
        self,
        timeout_ms: int,
    ) -> bool:
        del timeout_ms

        if (
            self._indice + 1
            >= len(self._paginas)
        ):
            return False

        self._indice += 1
        return True

    def capturar_screenshot(
        self,
        caminho: Path,
    ) -> None:
        caminho.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        caminho.write_bytes(b"PNG")

    def salvar_html(
        self,
        caminho: Path,
    ) -> None:
        caminho.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        caminho.write_text(
            "<html></html>",
            encoding="utf-8",
        )

    def fechar(self) -> None:
        pass


class AlertasFalsos:
    """Simula a entrega do alerta final."""

    def __init__(self) -> None:
        self.chamadas: list[
            dict[str, object]
        ] = []

    def enviar_alerta(
        self,
        **argumentos,
    ):
        self.chamadas.append(
            argumentos
        )

        return SimpleNamespace(
            sucesso=True,
            canal="email",
            erro=None,
        )


def test_pipeline_capstone_completo(
    tmp_path,
):
    """Executa A, B, C, D, E e F em uma cadeia completa."""

    entrada = criar_planilha_entrada(
        tmp_path / "entrada.xlsx"
    )

    resultado_a = executar_bot_entrada(
        entrada,
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
    )

    assert resultado_a.sucesso is True

    caminho_b = (
        tmp_path / "estoque.json"
    )

    resultado_b = (
        executar_bot_coleta_desktop(
            AutomacaoDesktopFalsa(),
            execution_id=EXECUTION_ID,
            correlation_id=CORRELATION_ID,
            task_id="task-b",
            predecessor=(
                "bot-a-entrada-capstone"
            ),
            predecessor_task_id=(
                "task-a"
            ),
            resultado_predecessor=(
                "pronto_para_processamento"
            ),
            configuracao=(
                ConfiguracaoColetaDesktop(
                    max_tentativas=1,
                    timeout_seconds=5,
                    backoff_seconds=0,
                    max_paginas=5,
                    caminho_artefato=caminho_b,
                    diretorio_screenshots=(
                        tmp_path / "desktop"
                    ),
                )
            ),
        )
    )

    assert resultado_b.sucesso is True
    assert resultado_b.total_registros == 3

    caminho_c = (
        tmp_path / "pedidos.json"
    )

    resultado_c = executar_bot_coleta_web(
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
        task_id="task-c",
        predecessor=(
            "bot-a-entrada-capstone"
        ),
        predecessor_task_id="task-a",
        resultado_predecessor=(
            "pronto_para_processamento"
        ),
        configuracao=(
            ConfiguracaoColetaWeb(
                portal_url=(
                    "http://portal-simulado"
                ),
                max_tentativas=1,
                timeout_seconds=5,
                backoff_seconds=0,
                max_paginas=5,
                headless=True,
                intervalo_paginas_seconds=0,
                caminho_artefato=caminho_c,
                diretorio_evidencias=(
                    tmp_path / "web"
                ),
            )
        ),
        fabrica_automacao=(
            AutomacaoWebFalsa
        ),
    )

    assert resultado_c.sucesso is True
    assert resultado_c.total_registros == 3

    caminho_d = (
        tmp_path / "consolidacao.json"
    )

    resultado_d = executar_bot_consolidacao(
        caminho_estoque=caminho_b,
        caminho_pedidos=caminho_c,
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
        task_id="task-d",
        predecessor_task_ids=(
            "task-b",
            "task-c",
        ),
        configuracao=(
            ConfiguracaoConsolidacao(
                caminho_artefato=caminho_d
            )
        ),
    )

    assert resultado_d.sucesso is True
    assert resultado_d.total_registros == 3

    classificador = Mock()

    classificador.classificar.return_value = (
        ResultadoDecisaoHibrida.de_ml(
            causa_provavel=(
                "falha_de_planejamento"
            ),
            confianca_ml=0.94,
            versao_modelo="capstone-1.0",
        )
    )

    caminho_e = (
        tmp_path / "classificados.json"
    )

    resultado_e = executar_bot_classificacao_ml(
        caminho_consolidacao=caminho_d,
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
        task_id="task-e",
        predecessor_task_id="task-d",
        configuracao=(
            ConfiguracaoClassificacaoML(
                caminho_artefato=caminho_e
            )
        ),
        classificador=classificador,
    )

    assert resultado_e.sucesso is True
    assert resultado_e.total_registros == 3
    assert resultado_e.total_ml == 3

    caminho_excel = (
        tmp_path / "relatorio.xlsx"
    )

    caminho_markdown = (
        tmp_path / "resumo.md"
    )

    alertas = AlertasFalsos()

    resultado_f = executar_bot_relatorio_capstone(
        caminho_classificacao=caminho_e,
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
        task_id="task-f",
        predecessor_task_id="task-e",
        configuracao=(
            ConfiguracaoRelatorioCapstone(
                caminho_excel=caminho_excel,
                caminho_markdown=(
                    caminho_markdown
                ),
            )
        ),
        sistema_alertas=alertas,
    )

    assert resultado_f.sucesso is True
    assert resultado_f.total_registros == 3
    assert caminho_excel.is_file()
    assert caminho_markdown.is_file()
    assert resultado_f.alerta_enviado is True

    artefato_b = (
        ArtefatoEstoqueDesktop.de_json(
            caminho_b.read_text(
                encoding="utf-8"
            )
        )
    )

    artefato_c = (
        ArtefatoPedidosFornecedor.de_json(
            caminho_c.read_text(
                encoding="utf-8"
            )
        )
    )

    artefato_d = (
        ArtefatoConsolidacao.de_json(
            caminho_d.read_text(
                encoding="utf-8"
            )
        )
    )

    artefato_e = (
        ArtefatoClassificacaoML.de_json(
            caminho_e.read_text(
                encoding="utf-8"
            )
        )
    )

    for auditoria in (
        artefato_b.auditoria,
        artefato_c.auditoria,
        artefato_d.auditoria,
        artefato_e.auditoria,
    ):
        assert (
            auditoria.execution_id
            == EXECUTION_ID
        )

        assert (
            auditoria.correlation_id
            == CORRELATION_ID
        )

    classificacoes = {
        registro
        .registro
        .classificacao_deterministica
        for registro
        in artefato_e.registros
    }

    assert classificacoes == {
        "REGULAR",
        "ESTOQUE_INDISPONIVEL",
        "ESTOQUE_INSUFICIENTE",
    }

    workbook = load_workbook(
        caminho_excel,
        read_only=True,
        data_only=True,
    )

    try:
        assert workbook[
            "Registros"
        ].max_row == 4
    finally:
        workbook.close()

    assert len(alertas.chamadas) == 1
    assert (
        alertas.chamadas[0]["severidade"]
        == "ERRO"
    )