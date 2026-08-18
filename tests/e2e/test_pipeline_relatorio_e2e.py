from datetime import datetime
import sys

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart

import gerar_relatorio

DIAS_INSPECAO = [
    ("Insp_15_06_2026", "15/06/2026"),
    ("Insp_16_06_2026", "16/06/2026"),
    ("Insp_17_06_2026", "17/06/2026"),
    ("Insp_18_06_2026", "18/06/2026"),
    ("Insp_19_06_2026", "19/06/2026"),
    ("Insp_22_06_2026", "22/06/2026"),
    ("Insp_23_06_2026", "23/06/2026"),
    ("Insp_24_06_2026", "24/06/2026"),
    ("Insp_25_06_2026", "25/06/2026"),
    ("Insp_26_06_2026", "26/06/2026"),
]

def criar_registro_base(lote_id, data):
    return {
        "lote_id": lote_id,
        "produto": "Produto A",
        "linha": "Linha 1",
        "turno": "Manhã",
        "status": "APROVADO",
        "responsavel": "Ana",
        "data": data,
        "observacao": "Inspeção concluída",
    }

def criar_registros_do_dia(indice, data):
    registro_1 = criar_registro_base(
        f"LOTE-{indice:03d}-A",
        data,
    )

    registro_2 = criar_registro_base(
        f"LOTE-{indice:03d}-B",
        data,
    )

    if indice == 1:
        # Confirma no fluxo completo a regressão NOK -> REPROVADO.
        registro_1["status"] = "NOK"
        registro_1["observacao"] = "Falha dimensional identificada"

    elif indice == 2:
        # Exercita a RN10: REPROVADO sem observação.
        registro_1["status"] = "REPROVADO"
        registro_1["observacao"] = None

    elif indice == 3:
        # Status não reconhecido -> Ambíguo.
        registro_1["status"] = "EM ANÁLISE"

    elif indice == 4:
        # Data inválida -> Erro de Entrada.
        registro_1["data"] = "2026-06-18"

    return [registro_1, registro_2]

def criar_planilha_entrada(caminho):
    lotes_referencia = []

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:

        for indice, (nome_aba, data) in enumerate(
            DIAS_INSPECAO,
            start=1,
        ):
            registros = criar_registros_do_dia(
                indice,
                data,
            )

            df = pd.DataFrame(registros)

            lotes_referencia.extend(df["lote_id"].tolist())

            df.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False,
                startrow=2,
            )

        base_referencia = pd.DataFrame(
            {
                "lote_id": lotes_referencia,
            }
        )

        base_referencia.to_excel(
            writer,
            sheet_name="Base_Referencia",
            index=False,
            startrow=1,
        )


@pytest.mark.e2e
def test_pipeline_completo_dez_dias_gera_relatorio_e_dashboard(
    tmp_path,
    monkeypatch,
):
    # Arrange
    entrada = tmp_path / "inspecao_lotes_10dias.xlsx"
    saida = tmp_path / "relatorio_conferencia_lotes.xlsx"

    criar_planilha_entrada(entrada)

    assert entrada.exists()

    workbook_entrada = load_workbook(
        entrada,
        read_only=True,
    )

    assert workbook_entrada.sheetnames == [
        "Insp_15_06_2026",
        "Insp_16_06_2026",
        "Insp_17_06_2026",
        "Insp_18_06_2026",
        "Insp_19_06_2026",
        "Insp_22_06_2026",
        "Insp_23_06_2026",
        "Insp_24_06_2026",
        "Insp_25_06_2026",
        "Insp_26_06_2026",
        "Base_Referencia",
    ]

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "gerar_relatorio.py",
            str(entrada),
            "--saida",
            str(saida),
        ],
    )

    # Act
    gerar_relatorio.main()

    # Assert
    assert saida.exists()
    assert saida.stat().st_size > 0

    workbook_saida = load_workbook(saida)

    # Assert
    assert saida.exists()
    assert saida.stat().st_size > 0

    workbook_saida = load_workbook(saida)

    abas_esperadas = [
        "Resumo",
        "Todos",
        "Válidos",
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Ranking de Regras",
        "Dicionário",
    ]

    assert workbook_saida.sheetnames == abas_esperadas

    resumo = workbook_saida["Resumo"]
    todos = workbook_saida["Todos"]
    validos = workbook_saida["Válidos"]
    divergencias = workbook_saida["Divergências"]
    ambiguos = workbook_saida["Ambíguos"]
    erros = workbook_saida["Erros de Entrada"]

    assert todos.max_row - 1 == 20

    assert validos.max_row - 1 == 17
    assert divergencias.max_row - 1 == 1
    assert ambiguos.max_row - 1 == 1
    assert erros.max_row - 1 == 1

    total_classificado = (
            (validos.max_row - 1)
            + (divergencias.max_row - 1)
            + (ambiguos.max_row - 1)
            + (erros.max_row - 1)
    )

    assert total_classificado == 20

    totais_dashboard = {
        resumo.cell(row=linha, column=1).value:
        resumo.cell(row=linha, column=2).value
        for linha in range(12, 16)
    }

    assert totais_dashboard == {
        "Válido": 17,
        "Divergência": 1,
        "Ambíguo": 1,
        "Erro de Entrada": 1,
    }

    percentuais = [
        resumo.cell(row=linha, column=3).value
        for linha in range(12, 16)
    ]

    assert sum(percentuais) == pytest.approx(1.0)

    datas_dashboard = [
        resumo.cell(row=linha, column=1).value
        for linha in range(30, 40)
    ]

    datas_esperadas = [
        datetime.strptime(data, "%d/%m/%Y")
        for _, data in DIAS_INSPECAO
    ]

    assert datas_dashboard == datas_esperadas

    assert len(resumo._charts) == 2
    assert isinstance(resumo._charts[0], DoughnutChart)
    assert isinstance(resumo._charts[1], LineChart)
