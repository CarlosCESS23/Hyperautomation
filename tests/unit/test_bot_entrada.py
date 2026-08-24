"""Testes unitários do Bot A de entrada."""

from unittest.mock import Mock

from openpyxl import Workbook

from src.bots.bot_entrada import (
    CAMPOS_OBRIGATORIOS_INSPECAO,
    StatusBotEntrada,
    executar_bot_entrada,
)


def criar_planilha_estrutural(
    caminho,
    *,
    remover_campo: str | None = None,
) -> None:
    """
    Cria uma planilha com estrutura válida.

    Não adicionamos registros porque o Bot A valida somente a estrutura.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)

    base_referencia = workbook.create_sheet(
        "Base_Referencia"
    )

    base_referencia.cell(
        row=1,
        column=1,
        value="Base de referência",
    )

    base_referencia.cell(
        row=2,
        column=1,
        value="lote_id",
    )

    inspecao = workbook.create_sheet(
        "Insp_15_06_2026"
    )

    inspecao.cell(
        row=1,
        column=1,
        value="Planilha de inspeção",
    )

    campos = sorted(CAMPOS_OBRIGATORIOS_INSPECAO)

    if remover_campo is not None:
        campos.remove(remover_campo)

    for coluna, campo in enumerate(campos, start=1):
        inspecao.cell(
            row=3,
            column=coluna,
            value=campo,
        )

    workbook.save(caminho)
    workbook.close()


def test_bot_entrada_prepara_parametros_para_bot_b(
    tmp_path,
):
    entrada = tmp_path / "entrada.xlsx"
    criar_planilha_estrutural(entrada)

    resultado = executar_bot_entrada(
        entrada,
        execution_id="exec-teste-001",
        correlation_id="corr-teste-001",
    )

    assert resultado.sucesso is True
    assert resultado.status is StatusBotEntrada.PRONTO
    assert resultado.parametros_bot_b is not None

    parametros = resultado.parametros_bot_b

    assert parametros.caminho_entrada == str(
        entrada.resolve()
    )
    assert parametros.execution_id == "exec-teste-001"
    assert parametros.correlation_id == "corr-teste-001"
    assert parametros.bot_predecessor == "bot-a-entrada"

    assert (
        parametros.resultado_predecessor
        == "pronto_para_conferencia"
    )


def test_arquivo_inexistente_produz_resultado_controlado(
    tmp_path,
):
    entrada = tmp_path / "nao_existe.xlsx"

    resultado = executar_bot_entrada(
        entrada,
        execution_id="exec-teste-002",
        correlation_id="corr-teste-002",
    )

    assert resultado.sucesso is False

    assert (
        resultado.status
        is StatusBotEntrada.ENTRADA_INVALIDA
    )

    assert "não encontrado" in resultado.mensagem
    assert resultado.parametros_bot_b is None

    # Mesmo uma execução inválida possui rastreabilidade.
    assert resultado.execution_id == "exec-teste-002"
    assert resultado.correlation_id == "corr-teste-002"


def test_planilha_sem_campo_obrigatorio_e_rejeitada(
    tmp_path,
):
    entrada = tmp_path / "sem_observacao.xlsx"

    criar_planilha_estrutural(
        entrada,
        remover_campo="observacao",
    )

    resultado = executar_bot_entrada(entrada)

    assert resultado.sucesso is False

    assert (
        resultado.status
        is StatusBotEntrada.ENTRADA_INVALIDA
    )

    assert "observacao" in resultado.mensagem
    assert resultado.parametros_bot_b is None


def test_bot_entrada_gera_identificadores_automaticamente(
    tmp_path,
):
    entrada = tmp_path / "entrada.xlsx"
    criar_planilha_estrutural(entrada)

    resultado = executar_bot_entrada(entrada)

    assert resultado.sucesso is True
    assert resultado.execution_id.startswith("exec-")
    assert resultado.correlation_id.startswith("corr-")

    assert (
        resultado.execution_id
        != resultado.correlation_id
    )


def test_bot_entrada_registra_inicio_e_conclusao(
    tmp_path,
):
    entrada = tmp_path / "entrada.xlsx"
    criar_planilha_estrutural(entrada)

    logger = Mock()

    resultado = executar_bot_entrada(
        entrada,
        execution_id="exec-log-001",
        correlation_id="corr-log-001",
        logger=logger,
    )

    assert resultado.sucesso is True
    assert logger.info.call_count == 2

    eventos = [
        chamada.kwargs["extra"]["evento"]
        for chamada in logger.info.call_args_list
    ]

    assert eventos == [
        "bot_entrada_iniciado",
        "bot_entrada_concluido",
    ]


def test_bot_entrada_nao_valida_conteudo_de_negocio(
    tmp_path,
):
    """
    A planilha sem registros continua estruturalmente válida.

    Quem analisará os registros e executará RN01–RN12 será o Bot B.
    """

    entrada = tmp_path / "sem_registros.xlsx"
    criar_planilha_estrutural(entrada)

    resultado = executar_bot_entrada(entrada)

    assert resultado.sucesso is True
    assert resultado.parametros_bot_b is not None