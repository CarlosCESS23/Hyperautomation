from datetime import datetime
from unittest.mock import Mock

from openpyxl import load_workbook

from gerar_relatorio import gerar_excel, ler_e_validar
from src.ml_decisions import AuditoriaDecisoesML
from src.ml_client import MLClient
from tests.integration.test_fluxo_relatorio_excel import criar_planilha_controlada


def test_fluxo_real_encaminha_decisoes_para_relatorio(
    tmp_path,
    monkeypatch,
):
    """
    Verifica se uma chamada real feita durante ler_e_validar()
    aparece na aba Decisões de ML.
    """

    entrada = tmp_path / "entrada_fluxo_real.xlsx"
    saida = tmp_path / "relatorio_fluxo_real.xlsx"

    criar_planilha_controlada(entrada)

    resposta_ml = {
        "classe": "revisar",
        "probabilidade": 0.75,
        "nivel_confianca": "revisar",
    }

    def classificar_simulado(
        self,
        *,
        status_raw: str,
        turno: str,
        tem_obs: bool,
    ):
        return resposta_ml

    # Evita chamada HTTP durante o teste.
    monkeypatch.setattr(
        MLClient,
        "classificar",
        classificar_simulado,
    )

    logger = Mock()

    auditoria = AuditoriaDecisoesML(
        logger=logger,
    )

    registros = ler_e_validar(
        entrada,
        auditoria_ml=auditoria,
    )

    gerar_excel(
        registros=registros,
        saida=saida,
        momento=datetime(2026, 6, 26, 12),
        decisoes_ml=auditoria.decisoes,
    )

    # A planilha controlada possui um registro ambíguo.
    assert len(auditoria.decisoes) == 1

    decisao = auditoria.decisoes[0]

    assert decisao.lote_id == "LOTE-AMBIGUO"
    assert decisao.classe_prevista == "revisar"
    assert decisao.probabilidade == 0.75
    assert decisao.nivel_confianca == "revisar"

    workbook = load_workbook(
        saida,
        data_only=True,
    )

    aba_ml = workbook["Decisões de ML"]

    # max_row inclui o cabeçalho.
    assert aba_ml.max_row - 1 == 1

    assert aba_ml.cell(2, 1).value == "LOTE-AMBIGUO"
    assert aba_ml.cell(2, 2).value == "revisar"
    assert aba_ml.cell(2, 3).value == 0.75
    assert aba_ml.cell(2, 4).value == "revisar"

    workbook.close()