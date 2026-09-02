"""Integração dos Bots E e F com o Maestro."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

from botcity.maestro import (
    AutomationTaskFinishStatus,
)
from openpyxl import load_workbook

from src.bots.bot_classificacao_ml import (
    ConfiguracaoClassificacaoML,
)
from src.bots.bot_classificacao_ml_maestro import (
    executar_tarefa_bot_e_capstone,
)
from src.bots.bot_relatorio_capstone import (
    ConfiguracaoRelatorioCapstone,
)
from src.bots.bot_relatorio_capstone_maestro import (
    executar_tarefa_bot_f_capstone,
)
from src.contratos_capstone import (
    ArtefatoClassificacaoML,
    ArtefatoConsolidacao,
    EnvelopeAuditoria,
    EstadoExecucao,
    FonteDados,
    RegistroConsolidado,
)
from src.decisao_hibrida import (
    ResultadoDecisaoHibrida,
)


EXECUTION_ID = "exec-capstone-001"
CORRELATION_ID = "corr-capstone-001"

TASK_D = "task-d-400"
TASK_E = "task-e-500"
TASK_F = "task-f-600"


def criar_consolidacao(
    caminho: Path,
) -> Path:
    """Cria o artefato que seria produzido pelo Bot D."""

    artefato = ArtefatoConsolidacao(
        auditoria=EnvelopeAuditoria(
            execution_id=EXECUTION_ID,
            correlation_id=CORRELATION_ID,
            bot_id="bot-d-consolidacao",
            task_id=TASK_D,
            estado=EstadoExecucao.CONCLUIDO,
            predecessor=(
                "bot-b-coleta-desktop"
                "+bot-c-coleta-web"
            ),
            predecessor_task_id=(
                "task-b-200,task-c-300"
            ),
            resultado_predecessor=(
                "fontes_avaliadas"
            ),
        ),
        registros=(
            RegistroConsolidado(
                lote_id="LOTE-001",
                produto=(
                    "Sensor de temperatura"
                ),
                quantidade_estoque=4,
                quantidade_pedida=7,
                status_estoque=(
                    "ESTOQUE_BAIXO"
                ),
                status_pedido="PENDENTE",
                classificacao_deterministica=(
                    "ESTOQUE_INSUFICIENTE"
                ),
                motivo=(
                    "Quantidade disponível menor "
                    "que a quantidade pedida."
                ),
                regras_aplicadas=("RD04",),
                fontes_disponiveis=(
                    FonteDados.DESKTOP,
                    FonteDados.WEB,
                ),
                modo_degradado=False,
            ),
        ),
    )

    caminho.write_text(
        artefato.para_json(),
        encoding="utf-8",
    )

    return caminho


def criar_baixador(
    origem: Path,
):
    """Simula o download de um artefato do Maestro."""

    def baixar(
        maestro,
        *,
        task_id_origem: str,
        nome_artefato: str,
        destino: str | Path,
        dias: int = 7,
    ) -> Path:
        del maestro
        del task_id_origem
        del nome_artefato
        del dias

        destino = Path(destino)

        destino.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        destino.write_bytes(
            origem.read_bytes()
        )

        return destino

    return baixar


class AlertasFalsos:
    """Canal controlado usado pelo Bot F."""

    def __init__(self) -> None:
        self.chamadas: list[
            dict[str, object]
        ] = []

    def enviar_alerta(
        self,
        **argumentos,
    ):
        self.chamadas.append(
            argumentos
        )

        return SimpleNamespace(
            sucesso=True,
            canal="email",
            erro=None,
        )


def test_bot_e_classifica_e_bot_f_gera_relatorio(
    tmp_path,
):
    """Executa a integração real entre E e F."""

    consolidacao = criar_consolidacao(
        tmp_path / "consolidacao.json"
    )

    saida_classificacao = (
        tmp_path
        / "classificados.json"
    )

    maestro_e = Mock()

    maestro_e.get_execution.return_value = (
        SimpleNamespace(
            task_id=TASK_E,
            parameters={
                "execution_id": EXECUTION_ID,
                "correlation_id": (
                    CORRELATION_ID
                ),
                "predecessor_task_id": TASK_D,
                "consolidacao_task_id": TASK_D,
                "consolidacao_artefato": (
                    "registros_consolidados.json"
                ),
            },
        )
    )

    maestro_e.get_task.return_value = (
        SimpleNamespace(
            status="FINISHED"
        )
    )

    classificador = Mock()

    classificador.classificar.return_value = (
        ResultadoDecisaoHibrida.de_ml(
            causa_provavel=(
                "falha_de_planejamento"
            ),
            confianca_ml=0.92,
            versao_modelo="capstone-1.0",
        )
    )

    servico = Mock()

    servico.agendar_relatorio.return_value = (
        SimpleNamespace(
            task_id=TASK_F
        )
    )

    codigo_e = executar_tarefa_bot_e_capstone(
        maestro=maestro_e,
        classificador=classificador,
        servico_orquestracao=servico,
        baixar=criar_baixador(
            consolidacao
        ),
        configuracao=(
            ConfiguracaoClassificacaoML(
                caminho_artefato=(
                    saida_classificacao
                )
            )
        ),
    )

    assert codigo_e == 0
    assert saida_classificacao.is_file()

    artefato_e = (
        ArtefatoClassificacaoML.de_json(
            saida_classificacao.read_text(
                encoding="utf-8"
            )
        )
    )

    assert artefato_e.total_registros == 1
    assert (
        artefato_e.registros[0]
        .origem_decisao
        == "ml"
    )

    servico.agendar_relatorio.assert_called_once_with(
        predecessor_task_id=TASK_E,
        resultado_predecessor="CONCLUIDO",
        execution_id=EXECUTION_ID,
        correlation_id=CORRELATION_ID,
        classificacao_task_id=TASK_E,
        classificacao_artefato=(
            "registros_classificados.json"
        ),
    )

    maestro_e.finish_task.assert_called_once()

    finalizacao_e = (
        maestro_e.finish_task
        .call_args
        .kwargs
    )

    assert (
        finalizacao_e["status"]
        == AutomationTaskFinishStatus.SUCCESS
    )

    maestro_f = Mock()

    maestro_f.get_execution.return_value = (
        SimpleNamespace(
            task_id=TASK_F,
            parameters={
                "execution_id": EXECUTION_ID,
                "correlation_id": (
                    CORRELATION_ID
                ),
                "predecessor_task_id": TASK_E,
                "classificacao_task_id": (
                    TASK_E
                ),
                "classificacao_artefato": (
                    "registros_classificados.json"
                ),
            },
        )
    )

    maestro_f.get_task.return_value = (
        SimpleNamespace(
            status="FINISHED"
        )
    )

    caminho_excel = (
        tmp_path
        / "relatorio_capstone.xlsx"
    )

    caminho_markdown = (
        tmp_path
        / "resumo_capstone.md"
    )

    alertas = AlertasFalsos()

    codigo_f = executar_tarefa_bot_f_capstone(
        maestro=maestro_f,
        sistema_alertas=alertas,
        baixar=criar_baixador(
            saida_classificacao
        ),
        configuracao=(
            ConfiguracaoRelatorioCapstone(
                caminho_excel=caminho_excel,
                caminho_markdown=(
                    caminho_markdown
                ),
            )
        ),
    )

    assert codigo_f == 0
    assert caminho_excel.is_file()
    assert caminho_markdown.is_file()

    workbook = load_workbook(
        caminho_excel,
        read_only=True,
        data_only=True,
    )

    try:
        assert workbook.sheetnames == [
            "Resumo",
            "Registros",
            "Decisoes ML",
            "Rastreabilidade",
        ]

        registros_ws = workbook[
            "Registros"
        ]

        assert registros_ws.max_row == 2
    finally:
        workbook.close()

    assert len(alertas.chamadas) == 1

    assert (
        alertas.chamadas[0]["anexo"]
        == caminho_excel
    )

    assert maestro_f.post_artifact.call_count == 2

    maestro_f.post_artifact.assert_any_call(
        task_id=TASK_F,
        artifact_name=(
            "relatorio_capstone.xlsx"
        ),
        filepath=str(caminho_excel),
    )

    maestro_f.post_artifact.assert_any_call(
        task_id=TASK_F,
        artifact_name=(
            "resumo_capstone.md"
        ),
        filepath=str(caminho_markdown),
    )

    finalizacao_f = (
        maestro_f.finish_task
        .call_args
        .kwargs
    )

    assert (
        finalizacao_f["status"]
        == AutomationTaskFinishStatus.SUCCESS
    )