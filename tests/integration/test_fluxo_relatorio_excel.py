"""Integração entre a entrada Excel, as regras de negócio e o relatório."""

from collections import Counter
from datetime import datetime

from openpyxl import Workbook, load_workbook
import pytest

from gerar_relatorio import gerar_excel, ler_e_validar


pytestmark = pytest.mark.integration

ABAS_ENTRADA = tuple(
    f"Insp_{dia:02d}_06_2026" for dia in (15, 16, 17, 18, 19, 22, 23, 24, 25, 26)
)
ABAS_RELATORIO = (
    "Resumo",
    "Todos",
    "Válidos",
    "Divergências",
    "Ambíguos",
    "Erros de Entrada",
    "Ranking de Regras",
    "Dicionário",
    "Decisões de ML",
)
CABECALHOS_ENTRADA = (
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
)
CABECALHOS_RELATORIO = (
    "Data de Referência",
    "Lote",
    "Produto",
    "Linha",
    "Turno",
    "Status",
    "Responsável",
    "Data da Inspeção",
    "Observação",
    "Classificação",
    "Motivo",
    "Ação Recomendada",
    "Regra Aplicada",
)
CLASSIFICACOES_ESPERADAS = {
    "LOTE-VALIDO": "Válido",
    "LOTE-DIVERGENTE": "Divergência",
    "LOTE-AMBIGUO": "Ambíguo",
    "LOTE-ERRO": "Erro de Entrada",
}


def criar_planilha_controlada(caminho):
    """Cria uma réplica mínima do contrato de entrada, inclusive sua base simulada."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    base = workbook.create_sheet("Base_Referencia")
    base.append(["Base de Referência simulada pelo teste"])
    base.append(["lote_id"])
    for lote in ("LOTE-VALIDO", "LOTE-AMBIGUO", "LOTE-ERRO"):
        base.append([lote])

    registros = (
        (
            "LOTE-VALIDO",
            "Produto A",
            "Linha 1",
            "Manhã",
            "APROVADO",
            "Ana",
            "15/06/2026",
            "Conferido",
        ),
        (
            "LOTE-DIVERGENTE",
            "Produto B",
            "Linha 2",
            "Tarde",
            "APROVADO",
            "Bruno",
            "15/06/2026",
            "Não cadastrado na referência",
        ),
        (
            "LOTE-AMBIGUO",
            "Produto C",
            "Linha 3",
            "Noite",
            "EM ANÁLISE",
            "Carla",
            "15/06/2026",
            "Aguardando decisão",
        ),
        (
            "LOTE-ERRO",
            None,
            "Linha 4",
            "Manhã",
            "APROVADO",
            "Diego",
            "15/06/2026",
            "Produto ausente",
        ),
    )

    for indice, nome_aba in enumerate(ABAS_ENTRADA):
        aba = workbook.create_sheet(nome_aba)
        aba.append([f"Inspeções de {nome_aba}"])
        aba.append([])
        aba.append(CABECALHOS_ENTRADA)
        if indice == 0:
            for registro in registros:
                aba.append(registro)

    workbook.save(caminho)


def valores_da_aba(workbook, nome_aba):
    """Retorna os registros de uma aba pública indexados pelo lote."""
    linhas = workbook[nome_aba].iter_rows(values_only=True)
    cabecalhos = next(linhas)
    return cabecalhos, {
        linha[1]: dict(zip(cabecalhos, linha)) for linha in linhas if linha[1] is not None
    }


def test_fluxo_da_planilha_ao_relatorio_preserva_e_classifica_registros(tmp_path):
    entrada = tmp_path / "inspecoes_controladas.xlsx"
    saida = tmp_path / "relatorio_conferencia.xlsx"
    criar_planilha_controlada(entrada)

    registros = ler_e_validar(entrada)
    dados_processados = gerar_excel(
        registros,
        saida,
        momento=datetime(2026, 6, 26, 12, 0, 0),
    )

    assert entrada.parent == tmp_path
    assert saida.parent == tmp_path
    assert saida.is_file()
    assert saida.stat().st_size > 0
    assert len(registros) == len(CLASSIFICACOES_ESPERADAS)
    assert len(dados_processados) == len(CLASSIFICACOES_ESPERADAS)

    relatorio = load_workbook(saida, data_only=False)
    assert tuple(relatorio.sheetnames) == ABAS_RELATORIO

    cabecalhos, todos = valores_da_aba(relatorio, "Todos")
    assert cabecalhos == CABECALHOS_RELATORIO
    assert len(todos) == len(CLASSIFICACOES_ESPERADAS)
    assert {
        lote: registro["Classificação"] for lote, registro in todos.items()
    } == CLASSIFICACOES_ESPERADAS
    assert todos["LOTE-DIVERGENTE"]["Regra Aplicada"] == "RN05"
    assert todos["LOTE-AMBIGUO"]["Regra Aplicada"] == "RN09"
    assert todos["LOTE-ERRO"]["Regra Aplicada"] == "RN02"

    contagens = Counter(registro["Classificação"] for registro in todos.values())
    resumo = relatorio["Resumo"]
    totais_resumo = {
        resumo.cell(linha, 1).value: resumo.cell(linha, 2).value
        for linha in range(12, 16)
    }
    assert resumo["A6"].value == len(CLASSIFICACOES_ESPERADAS)
    assert totais_resumo == dict(contagens)

    for classificacao, nome_aba in zip(
        ("Válido", "Divergência", "Ambíguo", "Erro de Entrada"),
        ABAS_RELATORIO[2:],
    ):
        cabecalhos_aba, registros_aba = valores_da_aba(relatorio, nome_aba)
        assert cabecalhos_aba == CABECALHOS_RELATORIO
        assert len(registros_aba) == contagens[classificacao] == 1
        assert all(
            registro["Classificação"] == classificacao
            for registro in registros_aba.values()
        )

    relatorio.close()
