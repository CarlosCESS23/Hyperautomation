"""Testes da rastreabilidade do pipeline híbrido no relatório Excel."""

from datetime import datetime

from openpyxl import load_workbook
import pytest

from gerar_relatorio import gerar_excel
from src.validacao_lotes import RegistroValidado


pytestmark = pytest.mark.integration


CABECALHOS_RASTREABILIDADE = (
    "Causa Provável",
    "Origem da Decisão",
    "Confiança ML",
    "Motivo do Fallback",
    "Versão do Modelo",
)


def criar_registro(
    *,
    lote: str,
    classificacao: str,
    causa_provavel: str = "",
    origem_decisao: str = "",
    confianca_ml: float | None = None,
    motivo_fallback: str = "",
    versao_modelo: str = "",
) -> RegistroValidado:
    """Cria um registro controlado para os testes do relatório."""

    return RegistroValidado(
        data_referencia="15/06/2026",
        lote=lote,
        produto="Produto Teste",
        linha="Linha 1",
        turno="Manhã",
        status="APROVADO",
        responsavel="Carlos",
        data_inspecao="15/06/2026",
        observacao="Observação controlada",
        classificacao=classificacao,
        motivo=(
            "Registro em conformidade"
            if classificacao == "Válido"
            else "Lote não encontrado na base de referência"
        ),
        acao_recomendada=(
            "Nenhuma ação necessária"
            if classificacao == "Válido"
            else "Conciliar com a base de referência ou com o processo"
        ),
        regra_aplicada=(
            ""
            if classificacao == "Válido"
            else "RN05"
        ),
        causa_provavel=causa_provavel,
        origem_decisao=origem_decisao,
        confianca_ml=confianca_ml,
        motivo_fallback=motivo_fallback,
        versao_modelo=versao_modelo,
    )


def ler_registros_da_aba(worksheet) -> list[dict]:
    """Converte as linhas do Excel em dicionários."""

    cabecalhos = [
        celula.value
        for celula in worksheet[1]
    ]

    return [
        dict(zip(cabecalhos, linha))
        for linha in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    ]


def test_relatorio_exibe_origem_confianca_e_fallback(tmp_path):
    saida = tmp_path / "relatorio_hibrido.xlsx"

    registros = [
        criar_registro(
            lote="LOTE-ML",
            classificacao="Divergência",
            causa_provavel="erro_codigo",
            origem_decisao="ml",
            confianca_ml=0.92,
            motivo_fallback="",
            versao_modelo="2.0.0-texto",
        ),
        criar_registro(
            lote="LOTE-FALLBACK",
            classificacao="Divergência",
            causa_provavel="nao_classificado",
            origem_decisao="fallback",
            confianca_ml=None,
            motivo_fallback="timeout",
            versao_modelo="",
        ),
        criar_registro(
            lote="LOTE-VALIDO",
            classificacao="Válido",
        ),
    ]

    gerar_excel(
        registros=registros,
        saida=saida,
        momento=datetime(2026, 6, 26, 12),
    )

    workbook = load_workbook(
        saida,
        data_only=False,
    )

    todos = workbook["Todos"]

    cabecalhos = tuple(
        celula.value
        for celula in todos[1]
    )

    for cabecalho in CABECALHOS_RASTREABILIDADE:
        assert cabecalho in cabecalhos

    linhas = {
        registro["Lote"]: registro
        for registro in ler_registros_da_aba(todos)
    }

    assert len(linhas) == len(registros)

    registro_ml = linhas["LOTE-ML"]

    assert registro_ml["Classificação"] == "Divergência"
    assert registro_ml["Regra Aplicada"] == "RN05"
    assert registro_ml["Causa Provável"] == "erro_codigo"
    assert registro_ml["Origem da Decisão"] == "ml"
    assert registro_ml["Confiança ML"] == pytest.approx(0.92)
    assert registro_ml["Motivo do Fallback"] is None
    assert registro_ml["Versão do Modelo"] == "2.0.0-texto"

    registro_fallback = linhas["LOTE-FALLBACK"]

    assert registro_fallback["Classificação"] == "Divergência"
    assert registro_fallback["Causa Provável"] == "nao_classificado"
    assert registro_fallback["Origem da Decisão"] == "fallback"
    assert registro_fallback["Confiança ML"] is None
    assert registro_fallback["Motivo do Fallback"] == "timeout"

    registro_valido = linhas["LOTE-VALIDO"]

    assert registro_valido["Classificação"] == "Válido"
    assert registro_valido["Origem da Decisão"] is None
    assert registro_valido["Confiança ML"] is None

    workbook.close()


def test_confianca_possui_formato_percentual(tmp_path):
    saida = tmp_path / "relatorio_formatado.xlsx"

    registros = [
        criar_registro(
            lote="LOTE-ML",
            classificacao="Divergência",
            causa_provavel="erro_codigo",
            origem_decisao="ml",
            confianca_ml=0.87,
            versao_modelo="2.0.0-texto",
        )
    ]

    gerar_excel(
        registros=registros,
        saida=saida,
        momento=datetime(2026, 6, 26, 12),
    )

    workbook = load_workbook(saida)
    worksheet = workbook["Divergências"]

    cabecalhos = {
        celula.value: celula.column
        for celula in worksheet[1]
    }

    coluna_confianca = cabecalhos["Confiança ML"]

    celula_confianca = worksheet.cell(
        row=2,
        column=coluna_confianca,
    )

    assert celula_confianca.value == pytest.approx(0.87)
    assert celula_confianca.number_format == "0.00%"

    workbook.close()


def test_todas_as_divergencias_sao_preservadas(tmp_path):
    saida = tmp_path / "relatorio_sem_perdas.xlsx"

    registros = [
        criar_registro(
            lote="LOTE-1",
            classificacao="Divergência",
            causa_provavel="erro_codigo",
            origem_decisao="ml",
            confianca_ml=0.91,
            versao_modelo="2.0.0-texto",
        ),
        criar_registro(
            lote="LOTE-2",
            classificacao="Divergência",
            causa_provavel="nao_classificado",
            origem_decisao="fallback",
            motivo_fallback="servico_indisponivel",
        ),
        criar_registro(
            lote="LOTE-3",
            classificacao="Divergência",
            causa_provavel="nao_classificado",
            origem_decisao="fallback",
            motivo_fallback="baixa_confianca",
        ),
    ]

    gerar_excel(
        registros=registros,
        saida=saida,
        momento=datetime(2026, 6, 26, 12),
    )

    workbook = load_workbook(saida)

    todos = workbook["Todos"]
    divergencias = workbook["Divergências"]

    # Desconta a primeira linha, que contém os cabeçalhos.
    assert todos.max_row - 1 == 3
    assert divergencias.max_row - 1 == 3

    linhas = ler_registros_da_aba(divergencias)

    assert {
        registro["Lote"]
        for registro in linhas
    } == {
        "LOTE-1",
        "LOTE-2",
        "LOTE-3",
    }

    assert {
        registro["Origem da Decisão"]
        for registro in linhas
    } == {
        "ml",
        "fallback",
    }

    workbook.close()

def test_aba_decisoes_ml_recebe_decisoes_do_fluxo_real(
    tmp_path,
):
    saida = (
        tmp_path
        / "relatorio_com_auditoria_hibrida.xlsx"
    )

    registros = [
        criar_registro(
            lote="LOTE-ML",
            classificacao="Divergência",
            causa_provavel="erro_codigo",
            origem_decisao="ml",
            confianca_ml=0.93,
            versao_modelo="2.0.0-texto",
        ),
        criar_registro(
            lote="LOTE-FALLBACK",
            classificacao="Divergência",
            causa_provavel="nao_classificado",
            origem_decisao="fallback",
            motivo_fallback="timeout",
        ),
        criar_registro(
            lote="LOTE-VALIDO",
            classificacao="Válido",
        ),
    ]

    gerar_excel(
        registros=registros,
        saida=saida,
        momento=datetime(
            2026,
            6,
            26,
            12,
        ),
    )

    workbook = load_workbook(
        saida,
        data_only=True,
    )

    worksheet = workbook[
        "Decisões de ML"
    ]

    # Somente os registros que passaram pela
    # decisão híbrida devem aparecer nessa aba.
    assert worksheet.max_row - 1 == 2

    linhas = ler_registros_da_aba(
        worksheet
    )

    registros_por_lote = {
        linha["Lote ID"]: linha
        for linha in linhas
    }

    assert set(registros_por_lote) == {
        "LOTE-ML",
        "LOTE-FALLBACK",
    }

    decisao_ml = registros_por_lote[
        "LOTE-ML"
    ]

    assert decisao_ml[
        "Origem da Decisão"
    ] == "ml"

    assert decisao_ml[
        "Causa Provável"
    ] == "erro_codigo"

    assert decisao_ml[
        "Confiança ML"
    ] == pytest.approx(0.93)

    decisao_fallback = registros_por_lote[
        "LOTE-FALLBACK"
    ]

    assert decisao_fallback[
        "Origem da Decisão"
    ] == "fallback"

    assert decisao_fallback[
        "Motivo do Fallback"
    ] == "timeout"

    workbook.close()