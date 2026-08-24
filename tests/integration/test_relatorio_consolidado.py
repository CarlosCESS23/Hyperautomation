"""Contrato integrado dos artefatos executivos da Aula 24."""

from datetime import datetime

from openpyxl import load_workbook
import pytest

from gerar_relatorio import gerar_excel, gerar_resumo_executivo, ler_e_validar
from src.operational_indicators import consolidar_indicadores
from tests.integration.test_fluxo_relatorio_excel import criar_planilha_controlada


pytestmark = pytest.mark.integration


def test_excel_e_markdown_compartilham_os_indicadores(tmp_path):
    entrada = tmp_path / "entrada.xlsx"
    excel = tmp_path / "relatorio_conferencia_lotes.xlsx"
    markdown = tmp_path / "resumo_executivo.md"
    criar_planilha_controlada(entrada)

    registros = ler_e_validar(entrada)
    indicadores = consolidar_indicadores(registros)
    gerar_excel(registros, excel, datetime(2026, 6, 26, 12), indicadores)
    gerar_resumo_executivo(indicadores, markdown)

    workbook = load_workbook(excel)
    assert workbook.sheetnames == [
        "Resumo", "Todos", "Válidos", "Divergências", "Ambíguos",
        "Erros de Entrada", "Ranking de Regras", "Dicionário",
        "Decisões de ML",
    ]
    resumo = workbook["Resumo"]
    assert resumo["B53"].value == indicadores.total_registros
    assert resumo["B58"].value == indicadores.regra_mais_acionada

    ranking = workbook["Ranking de Regras"]
    assert ranking["A2"].value == indicadores.regra_mais_acionada
    assert ranking["C2"].value == indicadores.ranking_regras[0].quantidade
    assert [ranking.cell(row, 3).value for row in range(2, ranking.max_row + 1)] == sorted(
        (regra.quantidade for regra in indicadores.ranking_regras), reverse=True
    )

    conteudo = markdown.read_text(encoding="utf-8")
    assert markdown.is_file()
    assert f"**{indicadores.total_registros} registros**" in conteudo
    assert f"**{indicadores.taxa_retrabalho:.1%}**" in conteudo
    assert f"**{indicadores.regra_mais_acionada}" in conteudo
    assert f"**{indicadores.ganho_estimado_horas:.2f} horas**" in conteudo
    workbook.close()
