"""Integração entre chamada ao classificador, log e relatório auditável."""

from datetime import datetime
import json
import logging

from openpyxl import load_workbook
from pythonjsonlogger import jsonlogger

from gerar_relatorio import gerar_excel, ler_e_validar
from src.ml_decisions import AuditoriaDecisoesML
from tests.integration.test_fluxo_relatorio_excel import criar_planilha_controlada


def test_uma_decisao_por_chamada_aparece_no_log_e_na_nona_aba(tmp_path):
    entrada = tmp_path / "entrada.xlsx"
    saida = tmp_path / "relatorio.xlsx"
    log = tmp_path / "ml.jsonl"
    criar_planilha_controlada(entrada)

    logger = logging.getLogger(f"teste.ml.{id(tmp_path)}")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    handler = logging.FileHandler(log, encoding="utf-8")
    handler.setFormatter(jsonlogger.JsonFormatter())
    logger.addHandler(handler)
    auditoria = AuditoriaDecisoesML(logger)

    respostas = (
        ("LOTE-ML-1", {"classe": "Válido", "probabilidade": 0.97, "nivel_confianca": "alta"}),
        ("LOTE-ML-2", {"classe": "Ambíguo", "probabilidade": 0.61, "nivel_confianca": "média"}),
        # Repetição intencional: decisões são chamadas, não lotes únicos.
        ("LOTE-ML-2", {"classe": "Divergência", "probabilidade": 0.88, "nivel_confianca": "alta"}),
    )
    for lote_id, resposta in respostas:
        assert auditoria.classificar(lote_id, lambda valor=resposta: valor) is resposta

    gerar_excel(
        ler_e_validar(entrada),
        saida,
        datetime(2026, 6, 26, 12),
        decisoes_ml=auditoria.decisoes,
    )
    handler.close()
    logger.removeHandler(handler)

    eventos = [json.loads(linha) for linha in log.read_text(encoding="utf-8").splitlines()]
    assert len(eventos) == len(respostas)
    assert all(evento["evento"] == "decisao_ml" for evento in eventos)
    assert all(
        {"lote_id", "classe_prevista", "probabilidade", "nivel_confianca", "latencia_ms"}
        <= evento.keys()
        for evento in eventos
    )

    workbook = load_workbook(saida, data_only=True)
    assert len(workbook.sheetnames) == 9
    assert workbook.sheetnames[-1] == "Decisões de ML"
    aba = workbook["Decisões de ML"]
    assert aba.max_row - 1 == len(respostas) == len(eventos)
    assert [aba.cell(linha, 1).value for linha in range(2, 5)] == [item[0] for item in respostas]
    assert [aba.cell(linha, 2).value for linha in range(2, 5)] == [
        item[1]["classe"] for item in respostas
    ]
    workbook.close()
