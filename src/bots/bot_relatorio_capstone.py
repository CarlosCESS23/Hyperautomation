"""Bot F: geração do relatório final do pipeline Capstone."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
import logging
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from src.contratos_capstone import (
    ArtefatoClassificacaoML,
    EstadoExecucao,
    RegistroClassificado,
)


LOGGER = logging.getLogger("botcity_permorfer")

BOT_ID = "bot-f-relatorio"


class PortaSistemaAlertas(Protocol):
    """Contrato necessário para enviar o relatório."""

    def enviar_alerta(
        self,
        *,
        severidade: str,
        mensagem: str,
        anexo: Path | None = None,
        contexto: dict[str, Any] | None = None,
    ) -> Any:
        """Envia o alerta pelo canal configurado."""


@dataclass(frozen=True)
class ConfiguracaoRelatorioCapstone:
    """Caminhos dos arquivos produzidos pelo Bot F."""

    caminho_excel: Path = Path(
        "reports/relatorio_capstone.xlsx"
    )

    caminho_markdown: Path = Path(
        "reports/resumo_capstone.md"
    )


@dataclass(frozen=True)
class ResumoRelatorioCapstone:
    """Contagens utilizadas no relatório e no alerta."""

    total_registros: int
    total_ml: int
    total_fallback: int
    total_degradados: int
    total_criticos: int
    classificacoes: dict[str, int]


@dataclass(frozen=True)
class ResultadoRelatorioCapstone:
    """Resultado controlado do Bot F."""

    sucesso: bool
    estado: EstadoExecucao
    total_registros: int
    caminho_excel: Path | None
    caminho_markdown: Path | None
    alerta_enviado: bool
    severidade_alerta: str
    erro_alerta: str | None = None
    erro: str | None = None


def carregar_classificacao(
    caminho: Path,
    *,
    execution_id: str,
    correlation_id: str,
) -> ArtefatoClassificacaoML:
    """Carrega e valida o artefato produzido pelo Bot E."""

    caminho = Path(caminho)

    if not caminho.is_file():
        raise FileNotFoundError(
            "Artefato classificado não encontrado: "
            f"{caminho}"
        )

    artefato = ArtefatoClassificacaoML.de_json(
        caminho.read_text(
            encoding="utf-8"
        )
    )

    if (
        artefato.auditoria.execution_id
        != execution_id
    ):
        raise ValueError(
            "execution_id do artefato classificado "
            "não pertence à execução atual"
        )

    if (
        artefato.auditoria.correlation_id
        != correlation_id
    ):
        raise ValueError(
            "correlation_id do artefato classificado "
            "não pertence à execução atual"
        )

    if artefato.auditoria.estado not in {
        EstadoExecucao.CONCLUIDO,
        EstadoExecucao.CONCLUIDO_DEGRADADO,
    }:
        raise ValueError(
            "Bot E não terminou com um "
            "estado processável"
        )

    return artefato


def consolidar_resumo(
    registros: tuple[
        RegistroClassificado,
        ...,
    ],
) -> ResumoRelatorioCapstone:
    """Consolida as informações apresentadas no relatório."""

    classificacoes = Counter(
        registro
        .registro
        .classificacao_deterministica
        for registro in registros
    )

    origens = Counter(
        registro.origem_decisao
        for registro in registros
    )

    classificacoes_criticas = {
        "ESTOQUE_INDISPONIVEL",
        "ESTOQUE_INSUFICIENTE",
        "SEM_REGISTRO_ESTOQUE",
        "ENTREGA_ATRASADA",
    }

    total_criticos = sum(
        quantidade
        for classificacao, quantidade
        in classificacoes.items()
        if classificacao
        in classificacoes_criticas
    )

    total_degradados = sum(
        1
        for registro in registros
        if registro.registro.modo_degradado
        or registro.origem_decisao
        == "fallback"
    )

    return ResumoRelatorioCapstone(
        total_registros=len(registros),
        total_ml=origens["ml"],
        total_fallback=origens["fallback"],
        total_degradados=total_degradados,
        total_criticos=total_criticos,
        classificacoes=dict(
            sorted(
                classificacoes.items()
            )
        ),
    )


def determinar_severidade(
    resumo: ResumoRelatorioCapstone,
) -> str:
    """Define a severidade operacional do encerramento."""

    if resumo.total_criticos > 0:
        return "ERRO"

    if (
        resumo.total_fallback > 0
        or resumo.total_degradados > 0
    ):
        return "AVISO"

    return "INFO"


def _estilizar_cabecalho(
    worksheet,
) -> None:
    """Aplica o estilo padrão ao cabeçalho da planilha."""

    preenchimento = PatternFill(
        "solid",
        fgColor="17365D",
    )

    fonte = Font(
        color="FFFFFF",
        bold=True,
    )

    for celula in worksheet[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def _adicionar_tabela(
    worksheet,
    nome: str,
) -> None:
    """Transforma o intervalo preenchido em tabela."""

    if (
        worksheet.max_row < 2
        or worksheet.max_column < 1
    ):
        return

    referencia = (
        f"A1:"
        f"{worksheet.cell(
            row=worksheet.max_row,
            column=worksheet.max_column,
        ).coordinate}"
    )

    tabela = Table(
        displayName=nome,
        ref=referencia,
    )

    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(tabela)


def _ajustar_larguras(
    worksheet,
) -> None:
    """Ajusta larguras sem deixar colunas exageradas."""

    for coluna in worksheet.columns:
        letra = coluna[0].column_letter

        maior = max(
            len(str(celula.value or ""))
            for celula in coluna
        )

        worksheet.column_dimensions[
            letra
        ].width = min(
            max(maior + 2, 12),
            45,
        )


def _linha_registro(
    registro: RegistroClassificado,
) -> tuple[object, ...]:
    """Converte um registro classificado em linha do Excel."""

    consolidado = registro.registro

    return (
        consolidado.lote_id,
        consolidado.produto,
        consolidado.quantidade_estoque,
        consolidado.quantidade_pedida,
        consolidado.status_estoque,
        consolidado.status_pedido,
        (
            consolidado
            .classificacao_deterministica
        ),
        consolidado.motivo,
        ", ".join(
            consolidado.regras_aplicadas
        ),
        ", ".join(
            fonte.value
            for fonte
            in consolidado.fontes_disponiveis
        ),
        (
            "SIM"
            if consolidado.modo_degradado
            else "NÃO"
        ),
        registro.causa_provavel,
        registro.origem_decisao,
        registro.confianca_ml,
        registro.motivo_fallback,
        registro.versao_modelo,
    )


def gerar_excel_capstone(
    artefato: ArtefatoClassificacaoML,
    resumo: ResumoRelatorioCapstone,
    caminho: Path,
) -> Path:
    """Gera o relatório Excel executivo do Capstone."""

    caminho = Path(caminho)

    caminho.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    resumo_ws = workbook.active
    resumo_ws.title = "Resumo"

    resumo_ws.append(
        (
            "Indicador",
            "Valor",
        )
    )

    indicadores = (
        (
            "Execution ID",
            artefato.auditoria.execution_id,
        ),
        (
            "Correlation ID",
            artefato.auditoria.correlation_id,
        ),
        (
            "Estado do Bot E",
            artefato.auditoria.estado.value,
        ),
        (
            "Gerado em",
            datetime.now(
                timezone.utc
            ).isoformat(),
        ),
        (
            "Total de registros",
            resumo.total_registros,
        ),
        (
            "Decisões por ML",
            resumo.total_ml,
        ),
        (
            "Decisões por fallback",
            resumo.total_fallback,
        ),
        (
            "Registros degradados",
            resumo.total_degradados,
        ),
        (
            "Registros críticos",
            resumo.total_criticos,
        ),
    )

    for indicador in indicadores:
        resumo_ws.append(indicador)

    linha_classificacoes = (
        resumo_ws.max_row + 3
    )

    resumo_ws.cell(
        linha_classificacoes,
        1,
        "Classificação",
    )

    resumo_ws.cell(
        linha_classificacoes,
        2,
        "Quantidade",
    )

    for numero, (
        classificacao,
        quantidade,
    ) in enumerate(
        resumo.classificacoes.items(),
        start=linha_classificacoes + 1,
    ):
        resumo_ws.cell(
            numero,
            1,
            classificacao,
        )

        resumo_ws.cell(
            numero,
            2,
            quantidade,
        )

    if resumo.classificacoes:
        grafico = BarChart()

        grafico.title = (
            "Distribuição das classificações"
        )

        grafico.y_axis.title = (
            "Quantidade"
        )

        grafico.x_axis.title = (
            "Classificação"
        )

        dados = Reference(
            resumo_ws,
            min_col=2,
            min_row=linha_classificacoes,
            max_row=(
                linha_classificacoes
                + len(resumo.classificacoes)
            ),
        )

        categorias = Reference(
            resumo_ws,
            min_col=1,
            min_row=(
                linha_classificacoes + 1
            ),
            max_row=(
                linha_classificacoes
                + len(resumo.classificacoes)
            ),
        )

        grafico.add_data(
            dados,
            titles_from_data=True,
        )

        grafico.set_categories(
            categorias
        )

        grafico.height = 8
        grafico.width = 16

        resumo_ws.add_chart(
            grafico,
            "D2",
        )

    registros_ws = workbook.create_sheet(
        "Registros"
    )

    registros_ws.append(
        (
            "Lote ID",
            "Produto",
            "Quantidade em Estoque",
            "Quantidade Pedida",
            "Status do Estoque",
            "Status do Pedido",
            "Classificação Determinística",
            "Motivo",
            "Regras Aplicadas",
            "Fontes Disponíveis",
            "Modo Degradado",
            "Causa Provável",
            "Origem da Decisão",
            "Confiança ML",
            "Motivo do Fallback",
            "Versão do Modelo",
        )
    )

    for registro in artefato.registros:
        registros_ws.append(
            _linha_registro(registro)
        )

    decisoes_ws = workbook.create_sheet(
        "Decisoes ML"
    )

    decisoes_ws.append(
        (
            "Lote ID",
            "Classificação Final",
            "Causa Provável",
            "Origem da Decisão",
            "Confiança ML",
            "Motivo do Fallback",
            "Versão do Modelo",
        )
    )

    for registro in artefato.registros:
        decisoes_ws.append(
            (
                registro.lote_id,
                registro.classificacao_final,
                registro.causa_provavel,
                registro.origem_decisao,
                registro.confianca_ml,
                registro.motivo_fallback,
                registro.versao_modelo,
            )
        )

    rastreabilidade_ws = (
        workbook.create_sheet(
            "Rastreabilidade"
        )
    )

    rastreabilidade_ws.append(
        (
            "Campo",
            "Valor",
        )
    )

    auditoria = artefato.auditoria

    rastreabilidade = (
        (
            "Schema Version",
            auditoria.schema_version,
        ),
        (
            "Execution ID",
            auditoria.execution_id,
        ),
        (
            "Correlation ID",
            auditoria.correlation_id,
        ),
        (
            "Bot de Origem",
            auditoria.bot_id,
        ),
        (
            "Task ID de Origem",
            auditoria.task_id,
        ),
        (
            "Estado",
            auditoria.estado.value,
        ),
        (
            "Predecessor",
            auditoria.predecessor,
        ),
        (
            "Predecessor Task ID",
            auditoria.predecessor_task_id,
        ),
        (
            "Resultado do Predecessor",
            auditoria.resultado_predecessor,
        ),
        (
            "Registrado em",
            auditoria.registrado_em.isoformat(),
        ),
    )

    for linha in rastreabilidade:
        rastreabilidade_ws.append(linha)

    for worksheet in workbook.worksheets:
        _estilizar_cabecalho(
            worksheet
        )

        _ajustar_larguras(
            worksheet
        )

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    _adicionar_tabela(
        registros_ws,
        "TabelaRegistrosCapstone",
    )

    _adicionar_tabela(
        decisoes_ws,
        "TabelaDecisoesCapstone",
    )

    borda = Border(
        bottom=Side(
            style="thin",
            color="CBD5E1",
        )
    )

    for linha in resumo_ws.iter_rows(
        min_row=2,
        max_row=resumo_ws.max_row,
        min_col=1,
        max_col=2,
    ):
        for celula in linha:
            celula.border = borda

    for worksheet in (
        registros_ws,
        decisoes_ws,
    ):
        for linha in range(
            2,
            worksheet.max_row + 1,
        ):
            celula_confianca = (
                worksheet.cell(
                    row=linha,
                    column=(
                        14
                        if worksheet.title
                        == "Registros"
                        else 5
                    ),
                )
            )

            if isinstance(
                celula_confianca.value,
                (int, float),
            ):
                celula_confianca.number_format = (
                    "0.00%"
                )

    temporario = caminho.with_suffix(
        ".tmp.xlsx"
    )

    workbook.save(
        temporario
    )

    temporario.replace(
        caminho
    )

    return caminho


def gerar_markdown_capstone(
    artefato: ArtefatoClassificacaoML,
    resumo: ResumoRelatorioCapstone,
    caminho: Path,
) -> Path:
    """Gera um resumo textual para apresentação."""

    caminho = Path(caminho)

    caminho.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    linhas = [
        "# Resumo Executivo — Capstone",
        "",
        f"- Execution ID: `{artefato.auditoria.execution_id}`",
        f"- Correlation ID: `{artefato.auditoria.correlation_id}`",
        f"- Estado: **{artefato.auditoria.estado.value}**",
        f"- Total processado: **{resumo.total_registros}**",
        f"- Decisões por ML: **{resumo.total_ml}**",
        f"- Decisões por fallback: **{resumo.total_fallback}**",
        f"- Registros degradados: **{resumo.total_degradados}**",
        f"- Registros críticos: **{resumo.total_criticos}**",
        "",
        "## Classificações",
        "",
    ]

    if resumo.classificacoes:
        for (
            classificacao,
            quantidade,
        ) in resumo.classificacoes.items():
            linhas.append(
                f"- {classificacao}: "
                f"**{quantidade}**"
            )
    else:
        linhas.append(
            "- Nenhum registro classificado."
        )

    linhas.extend(
        [
            "",
            "## Observação",
            "",
            (
                "A classificação final é definida "
                "pelas regras determinísticas. "
                "O ML sugere somente a causa provável."
            ),
            "",
        ]
    )

    caminho.write_text(
        "\n".join(linhas),
        encoding="utf-8",
    )

    return caminho


def montar_mensagem_alerta(
    *,
    resumo: ResumoRelatorioCapstone,
    execution_id: str,
    correlation_id: str,
    caminho_excel: Path,
) -> str:
    """Monta uma mensagem resumida para Telegram e Email."""

    linhas = [
        "Pipeline Capstone concluído.",
        "",
        f"Total processado: {resumo.total_registros}",
        f"Decisões por ML: {resumo.total_ml}",
        (
            "Decisões por fallback: "
            f"{resumo.total_fallback}"
        ),
        (
            "Registros degradados: "
            f"{resumo.total_degradados}"
        ),
        (
            "Registros críticos: "
            f"{resumo.total_criticos}"
        ),
        "",
        "Classificações:",
    ]

    for (
        classificacao,
        quantidade,
    ) in resumo.classificacoes.items():
        linhas.append(
            f"- {classificacao}: {quantidade}"
        )

    linhas.extend(
        [
            "",
            f"Execution ID: {execution_id}",
            f"Correlation ID: {correlation_id}",
            f"Relatório: {caminho_excel.name}",
        ]
    )

    return "\n".join(linhas)


def executar_bot_relatorio_capstone(
    *,
    caminho_classificacao: Path,
    execution_id: str,
    correlation_id: str,
    task_id: str,
    predecessor_task_id: str,
    configuracao: (
        ConfiguracaoRelatorioCapstone | None
    ) = None,
    sistema_alertas: (
        PortaSistemaAlertas | None
    ) = None,
    logger: logging.Logger | None = None,
) -> ResultadoRelatorioCapstone:
    """Gera o relatório final e encerra a cadeia."""

    del task_id
    del predecessor_task_id

    configuracao = (
        configuracao
        or ConfiguracaoRelatorioCapstone()
    )

    logger = logger or LOGGER

    try:
        artefato = carregar_classificacao(
            caminho_classificacao,
            execution_id=execution_id,
            correlation_id=correlation_id,
        )

        resumo = consolidar_resumo(
            artefato.registros
        )

        caminho_excel = gerar_excel_capstone(
            artefato,
            resumo,
            configuracao.caminho_excel,
        )

        caminho_markdown = (
            gerar_markdown_capstone(
                artefato,
                resumo,
                configuracao.caminho_markdown,
            )
        )

        severidade = determinar_severidade(
            resumo
        )

        alerta_enviado = False
        erro_alerta = None

        if sistema_alertas is not None:
            try:
                resultado_alerta = (
                    sistema_alertas
                    .enviar_alerta(
                        severidade=severidade,
                        mensagem=(
                            montar_mensagem_alerta(
                                resumo=resumo,
                                execution_id=(
                                    execution_id
                                ),
                                correlation_id=(
                                    correlation_id
                                ),
                                caminho_excel=(
                                    caminho_excel
                                ),
                            )
                        ),
                        anexo=caminho_excel,
                        contexto={
                            "bot_id": BOT_ID,
                            "execution_id": (
                                execution_id
                            ),
                            "correlation_id": (
                                correlation_id
                            ),
                            "total_registros": (
                                resumo.total_registros
                            ),
                            "total_ml": (
                                resumo.total_ml
                            ),
                            "total_fallback": (
                                resumo.total_fallback
                            ),
                            "total_criticos": (
                                resumo.total_criticos
                            ),
                        },
                    )
                )

                alerta_enviado = bool(
                    getattr(
                        resultado_alerta,
                        "sucesso",
                        False,
                    )
                )

                if not alerta_enviado:
                    erro_alerta = str(
                        getattr(
                            resultado_alerta,
                            "erro",
                            "alerta não entregue",
                        )
                    )

            except Exception as erro:
                erro_alerta = str(erro)

                logger.exception(
                    "alerta_relatorio_capstone_falhou",
                    extra={
                        "evento": (
                            "alerta_relatorio_capstone_falhou"
                        ),
                        "bot_id": BOT_ID,
                        "execution_id": (
                            execution_id
                        ),
                        "correlation_id": (
                            correlation_id
                        ),
                        "erro": erro_alerta,
                    },
                )

        estado = artefato.auditoria.estado

        if (
            sistema_alertas is not None
            and not alerta_enviado
        ):
            estado = (
                EstadoExecucao
                .CONCLUIDO_DEGRADADO
            )

        logger.info(
            "cadeia_capstone_encerrada",
            extra={
                "evento": (
                    "cadeia_capstone_encerrada"
                ),
                "bot_id": BOT_ID,
                "execution_id": execution_id,
                "correlation_id": (
                    correlation_id
                ),
                "estado": estado.value,
                "total_registros": (
                    resumo.total_registros
                ),
                "total_ml": resumo.total_ml,
                "total_fallback": (
                    resumo.total_fallback
                ),
                "total_criticos": (
                    resumo.total_criticos
                ),
                "alerta_enviado": (
                    alerta_enviado
                ),
                "erro_alerta": (
                    erro_alerta
                ),
                "caminho_excel": (
                    str(caminho_excel)
                ),
                "caminho_markdown": (
                    str(caminho_markdown)
                ),
            },
        )

        return ResultadoRelatorioCapstone(
            sucesso=True,
            estado=estado,
            total_registros=(
                resumo.total_registros
            ),
            caminho_excel=caminho_excel,
            caminho_markdown=(
                caminho_markdown
            ),
            alerta_enviado=(
                alerta_enviado
            ),
            severidade_alerta=(
                severidade
            ),
            erro_alerta=erro_alerta,
        )

    except Exception as erro:
        logger.exception(
            "relatorio_capstone_falhou",
            extra={
                "evento": (
                    "relatorio_capstone_falhou"
                ),
                "bot_id": BOT_ID,
                "execution_id": execution_id,
                "correlation_id": (
                    correlation_id
                ),
                "erro": str(erro),
            },
        )

        return ResultadoRelatorioCapstone(
            sucesso=False,
            estado=EstadoExecucao.FALHOU,
            total_registros=0,
            caminho_excel=None,
            caminho_markdown=None,
            alerta_enviado=False,
            severidade_alerta="ERRO",
            erro=str(erro),
        )