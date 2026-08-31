"""Bot A: recebe e prepara a entrada para o Bot B."""

from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass
from enum import Enum
import json
import logging
from pathlib import Path
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


LOGGER = logging.getLogger("botcity_permorfer")

BOT_ID = "bot-a-entrada"

ABA_BASE_REFERENCIA = "Base_Referencia"

CAMPOS_OBRIGATORIOS_INSPECAO = frozenset(
    {
        "lote_id",
        "produto",
        "linha",
        "turno",
        "status",
        "responsavel",
        "data",
        "observacao",
    }
)


class StatusBotEntrada(str, Enum):
    """Resultados controlados do Bot A."""

    PRONTO = "pronto_para_conferencia"
    ENTRADA_INVALIDA = "entrada_invalida"


@dataclass(frozen=True)
class ParametrosBotConferencia:
    """Parâmetros preparados para o futuro Bot B."""

    caminho_entrada: str
    execution_id: str
    correlation_id: str
    bot_predecessor: str
    resultado_predecessor: str

    def to_dict(self) -> dict[str, str]:
        return asdict(self)


@dataclass(frozen=True)
class ResultadoBotEntrada:
    """Resultado controlado da execução do Bot A."""

    sucesso: bool
    status: StatusBotEntrada
    mensagem: str
    execution_id: str
    correlation_id: str
    parametros_bot_b: ParametrosBotConferencia | None = None

    def to_dict(self) -> dict[str, Any]:
        """Converte o resultado para log, JSON ou orquestração."""

        return {
            "sucesso": self.sucesso,
            "status": self.status.value,
            "mensagem": self.mensagem,
            "execution_id": self.execution_id,
            "correlation_id": self.correlation_id,
            "parametros_bot_b": (
                self.parametros_bot_b.to_dict()
                if self.parametros_bot_b is not None
                else None
            ),
        }


def gerar_execution_id() -> str:
    """Cria um identificador único para a execução completa."""

    return f"exec-{uuid4()}"


def gerar_correlation_id() -> str:
    """Cria um identificador para relacionar os bots da cadeia."""

    return f"corr-{uuid4()}"


def _cabecalhos_da_linha(
    worksheet,
    numero_linha: int,
) -> set[str]:
    """Obtém os cabeçalhos preenchidos de uma linha da planilha."""

    return {
        str(celula.value).strip()
        for celula in worksheet[numero_linha]
        if celula.value is not None
        and str(celula.value).strip()
    }


def validar_estrutura_planilha(
    caminho: Path,
) -> tuple[bool, str]:
    """
    Valida somente a estrutura mínima da entrada.

    Esta função não executa regras de negócio e não valida o conteúdo dos
    registros. Essa responsabilidade pertence ao Bot B.
    """

    if not caminho.exists():
        return (
            False,
            f"Arquivo de entrada não encontrado: {caminho}",
        )

    if not caminho.is_file():
        return (
            False,
            f"O caminho informado não é um arquivo: {caminho}",
        )

    if caminho.suffix.lower() != ".xlsx":
        return (
            False,
            "O arquivo de entrada deve possuir a extensão .xlsx",
        )

    try:
        workbook = load_workbook(
            caminho,
            read_only=True,
            data_only=True,
        )
    except (
        OSError,
        BadZipFile,
        InvalidFileException,
    ) as erro:
        return (
            False,
            f"Não foi possível abrir a planilha: {erro}",
        )

    try:
        if ABA_BASE_REFERENCIA not in workbook.sheetnames:
            return (
                False,
                "A aba Base_Referencia não foi encontrada",
            )

        abas_inspecao = [
            nome
            for nome in workbook.sheetnames
            if nome.startswith("Insp_")
        ]

        if not abas_inspecao:
            return (
                False,
                "Nenhuma aba de inspeção foi encontrada",
            )

        base_referencia = workbook[ABA_BASE_REFERENCIA]

        cabecalhos_base = _cabecalhos_da_linha(
            base_referencia,
            numero_linha=2,
        )

        if "lote_id" not in cabecalhos_base:
            return (
                False,
                "A aba Base_Referencia deve possuir a coluna lote_id",
            )

        for nome_aba in abas_inspecao:
            worksheet = workbook[nome_aba]

            cabecalhos = _cabecalhos_da_linha(
                worksheet,
                numero_linha=3,
            )

            campos_faltantes = (
                CAMPOS_OBRIGATORIOS_INSPECAO
                - cabecalhos
            )

            if campos_faltantes:
                campos_formatados = ", ".join(
                    sorted(campos_faltantes)
                )

                return (
                    False,
                    (
                        f"A aba {nome_aba} não possui os campos: "
                        f"{campos_formatados}"
                    ),
                )

        return (
            True,
            "Estrutura da planilha validada com sucesso",
        )
    finally:
        workbook.close()


def executar_bot_entrada(
    caminho_entrada: str | Path,
    *,
    execution_id: str | None = None,
    correlation_id: str | None = None,
    logger: logging.Logger | None = None,
) -> ResultadoBotEntrada:
    """
    Executa o Bot A isoladamente.

    Os identificadores podem ser injetados nos testes. Quando não forem
    informados, novos UUIDs serão gerados.
    """

    logger = logger or LOGGER

    execution_id = (
        execution_id.strip()
        if execution_id and execution_id.strip()
        else gerar_execution_id()
    )

    correlation_id = (
        correlation_id.strip()
        if correlation_id and correlation_id.strip()
        else gerar_correlation_id()
    )

    caminho = Path(caminho_entrada)

    logger.info(
        "bot_entrada_iniciado",
        extra={
            "evento": "bot_entrada_iniciado",
            "bot_id": BOT_ID,
            "execution_id": execution_id,
            "correlation_id": correlation_id,
            "caminho_entrada": str(caminho),
        },
    )

    estrutura_valida, mensagem = validar_estrutura_planilha(
        caminho
    )

    if not estrutura_valida:
        resultado = ResultadoBotEntrada(
            sucesso=False,
            status=StatusBotEntrada.ENTRADA_INVALIDA,
            mensagem=mensagem,
            execution_id=execution_id,
            correlation_id=correlation_id,
            parametros_bot_b=None,
        )

        logger.warning(
            "bot_entrada_rejeitado",
            extra={
                "evento": "bot_entrada_rejeitado",
                "bot_id": BOT_ID,
                **resultado.to_dict(),
            },
        )

        return resultado

    parametros_bot_b = ParametrosBotConferencia(
        caminho_entrada=str(caminho.resolve()),
        execution_id=execution_id,
        correlation_id=correlation_id,
        bot_predecessor=BOT_ID,
        resultado_predecessor=StatusBotEntrada.PRONTO.value,
    )

    resultado = ResultadoBotEntrada(
        sucesso=True,
        status=StatusBotEntrada.PRONTO,
        mensagem=mensagem,
        execution_id=execution_id,
        correlation_id=correlation_id,
        parametros_bot_b=parametros_bot_b,
    )

    logger.info(
        "bot_entrada_concluido",
        extra={
            "evento": "bot_entrada_concluido",
            "bot_id": BOT_ID,
            **resultado.to_dict(),
        },
    )

    return resultado


def main(argv: list[str] | None = None) -> int:
    """Ponto de entrada para executar o Bot A pelo terminal."""

    parser = argparse.ArgumentParser(
        description="Valida e prepara uma planilha para o Bot B",
    )

    parser.add_argument(
        "entrada",
        help="Caminho da planilha .xlsx",
    )

    args = parser.parse_args(argv)

    resultado = executar_bot_entrada(args.entrada)

    print(
        json.dumps(
            resultado.to_dict(),
            ensure_ascii=False,
            indent=2,
        )
    )

    return 0 if resultado.sucesso else 2


if __name__ == "__main__":
    raise SystemExit(main())