"""Gera o relatório executivo de conferência de lotes (RN01 a RN12).

Uso:
    python gerar_relatorio.py caminho/inspecao_lotes_10dias.xlsx

Sem argumento, o script procura o arquivo em ``data/input`` e em Downloads.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.validacao_lotes import (
    CLASSIFICACOES,
    RegistroValidado,
    texto,
    validar_registro,
)

CORES = {
    "Válido": "22C55E",
    "Divergência": "F59E0B",
    "Ambíguo": "8B5CF6",
    "Erro de Entrada": "EF4444",
}
ABAS = {
    "Válido": "Válidos",
    "Divergência": "Divergências",
    "Ambíguo": "Ambíguos",
    "Erro de Entrada": "Erros de Entrada",
}
ABAS_DIARIAS_ESPERADAS = tuple(
    f"Insp_{dia:02d}_06_2026" for dia in (15, 16, 17, 18, 19, 22, 23, 24, 25, 26)
)
TOTAIS_GABARITO = {
    "Válido": 150,
    "Divergência": 50,
    "Ambíguo": 20,
    "Erro de Entrada": 30,
}


def ler_e_validar(caminho: Path) -> list[RegistroValidado]:
    planilha = pd.ExcelFile(caminho)
    abas_diarias = sorted(
        (aba for aba in planilha.sheet_names if aba.startswith("Insp_")),
        key=lambda aba: datetime.strptime(aba, "Insp_%d_%m_%Y"),
    )
    faltantes = sorted(set(ABAS_DIARIAS_ESPERADAS) - set(abas_diarias))
    inesperadas = sorted(set(abas_diarias) - set(ABAS_DIARIAS_ESPERADAS))
    if faltantes or inesperadas:
        raise ValueError(
            "As abas diárias não correspondem aos 10 dias esperados. "
            f"Faltantes: {faltantes or 'nenhuma'}; "
            f"inesperadas: {inesperadas or 'nenhuma'}."
        )
    if "Base_Referencia" not in planilha.sheet_names:
        raise ValueError("A aba obrigatória Base_Referencia não foi encontrada.")

    referencia = pd.read_excel(caminho, sheet_name="Base_Referencia", header=1)
    lotes_referencia = {texto(valor) for valor in referencia["lote_id"] if texto(valor)}
    resultados: list[RegistroValidado] = []

    for aba in abas_diarias:
        dados = pd.read_excel(caminho, sheet_name=aba, header=2)
        # Remove somente rodapés/linhas sem os oito campos do registro.
        dados = dados[dados[["lote_id", "produto", "linha", "turno", "status", "responsavel", "data", "observacao"]].notna().any(axis=1)]
        dados = dados[~dados["lote_id"].fillna("").astype(str).str.startswith("Total de registros:")]
        data_referencia = datetime.strptime(aba, "Insp_%d_%m_%Y").strftime("%d/%m/%Y")

        # RN11 é deliberadamente reiniciada a cada aba/dia.
        totais = Counter(texto(valor) for valor in dados["lote_id"] if texto(valor))
        vistas: Counter[str] = Counter()
        for _, registro in dados.iterrows():
            lote = texto(registro.get("lote_id"))
            if lote and totais[lote] > 1:
                vistas[lote] += 1
                ocorrencia = vistas[lote]
            else:
                ocorrencia = 1
            resultados.append(
                validar_registro(registro, data_referencia, lotes_referencia, ocorrencia)
            )
    return resultados


def estilizar_tabela(ws, nome_tabela: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        tabela = Table(displayName=nome_tabela, ref=ws.dimensions)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(tabela)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="17365D")
        celula.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for coluna in ws.columns:
        valores = [len(str(c.value or "")) for c in coluna[:80]]
        largura = min(max(max(valores, default=8) + 2, 11), 48)
        ws.column_dimensions[get_column_letter(coluna[0].column)].width = largura
    for row in ws.iter_rows(min_row=2):
        for celula in row:
            celula.alignment = Alignment(vertical="top", wrap_text=True)


def montar_resumo(ws, df: pd.DataFrame, momento: datetime) -> None:
    azul, azul_claro, branco = "17365D", "D9EAF7", "FFFFFF"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N2")
    ws["A1"] = "CONFERÊNCIA DE LOTES · PAINEL EXECUTIVO"
    ws["A1"].font = Font(size=22, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws["A3"] = "Período analisado"
    ws["B3"] = f"{df['Data de Referência'].min()} a {df['Data de Referência'].max()}"
    ws["F3"] = "Atualizado em"
    ws["G3"] = momento.strftime("%d/%m/%Y %H:%M:%S")

    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    total = len(df)
    cards = [("TOTAL", total, "17365D")] + [
        (nome.upper(), int(contagens[nome]), CORES[nome]) for nome in CLASSIFICACOES
    ]
    for indice, (titulo, valor, cor) in enumerate(cards):
        coluna = 1 + indice * 3
        ws.merge_cells(start_row=5, start_column=coluna, end_row=5, end_column=coluna + 1)
        ws.merge_cells(start_row=6, start_column=coluna, end_row=6, end_column=coluna + 1)
        ws.merge_cells(start_row=7, start_column=coluna, end_row=7, end_column=coluna + 1)
        topo = ws.cell(5, coluna, titulo)
        numero = ws.cell(6, coluna, valor)
        for row in range(5, 8):
            for col in range(coluna, coluna + 2):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=cor)
        topo.font = Font(bold=True, color=branco, size=10)
        numero.font = Font(bold=True, color=branco, size=23)
        topo.alignment = numero.alignment = Alignment(horizontal="center", vertical="center")
        if titulo != "TOTAL":
            ws.cell(7, coluna).value = int(valor) / total if total else 0
            ws.cell(7, coluna).number_format = "0.0%"
            ws.cell(7, coluna).font = Font(bold=True, color=branco)
            ws.cell(7, coluna).alignment = Alignment(horizontal="center")

    ws["A10"] = "Distribuição por classificação"
    ws["A10"].font = Font(size=14, bold=True, color=azul)
    ws["A11"], ws["B11"], ws["C11"] = "Classificação", "Quantidade", "%"
    for i, nome in enumerate(CLASSIFICACOES, start=12):
        ws.cell(i, 1, nome)
        ws.cell(i, 2, int(contagens[nome]))
        ws.cell(i, 3, int(contagens[nome]) / total if total else 0)
        ws.cell(i, 3).number_format = "0.0%"

    rosca = DoughnutChart()
    rosca.title = "Distribuição dos 250 registros"
    rosca.holeSize = 58
    rosca.height, rosca.width = 8.2, 12.5
    rosca.add_data(Reference(ws, min_col=2, min_row=11, max_row=15), titles_from_data=True)
    rosca.set_categories(Reference(ws, min_col=1, min_row=12, max_row=15))
    rosca.dataLabels = DataLabelList()
    rosca.dataLabels.showPercent = True
    rosca.dataLabels.showLeaderLines = True
    rosca.series[0].data_points = [DataPoint(idx=i, spPr=None) for i in range(4)]
    for ponto, nome in zip(rosca.series[0].data_points, CLASSIFICACOES):
        ponto.graphicalProperties.solidFill = CORES[nome]
    ws.add_chart(rosca, "E10")

    ws["A28"] = "Evolução diária dos registros que exigem ação"
    ws["A28"].font = Font(size=14, bold=True, color=azul)
    ws["A29"], ws["B29"], ws["C29"], ws["D29"] = (
        "Data", "Divergências", "Ambíguos", "Total de problemas"
    )
    diario = (
        df.assign(_n=1)
        .pivot_table(index="Data de Referência", columns="Classificação", values="_n", aggfunc="sum", fill_value=0)
        .reindex(columns=CLASSIFICACOES, fill_value=0)
    )
    diario.index = pd.to_datetime(diario.index, format="%d/%m/%Y")
    diario = diario.sort_index()
    for row, (data, valores) in enumerate(diario.iterrows(), start=30):
        ws.cell(row, 1, data.to_pydatetime())
        ws.cell(row, 1).number_format = "dd/mm/yyyy"
        ws.cell(row, 2, int(valores["Divergência"]))
        ws.cell(row, 3, int(valores["Ambíguo"]))
        ws.cell(row, 4, int(valores["Divergência"] + valores["Ambíguo"] + valores["Erro de Entrada"]))

    linha = LineChart()
    linha.title = "Evolução dos registros"
    linha.y_axis.title = "Quantidade"
    linha.x_axis.title = "Dia da inspeção"
    linha.style = 13
    linha.height, linha.width = 9, 22
    linha.add_data(Reference(ws, min_col=2, max_col=4, min_row=29, max_row=29 + len(diario)), titles_from_data=True)
    linha.set_categories(Reference(ws, min_col=1, min_row=30, max_row=29 + len(diario)))
    for serie, cor in zip(linha.series, (CORES["Divergência"], CORES["Ambíguo"], CORES["Erro de Entrada"])):
        serie.graphicalProperties.line.solidFill = cor
        serie.graphicalProperties.line.width = 28575
        serie.marker.symbol = "circle"
        serie.marker.size = 7
    ws.add_chart(linha, "E28")

    ws["A43"] = "LEITURA PARA DECISÃO"
    ws["A44"] = "Corrigir na origem"
    ws["B44"] = f"{contagens['Erro de Entrada']} registros"
    ws["A45"] = "Conciliar com base/processo"
    ws["B45"] = f"{contagens['Divergência']} registros"
    ws["A46"] = "Decisão humana"
    ws["B46"] = f"{contagens['Ambíguo']} registros"
    ws["A48"] = "Nota: duplicidades são avaliadas separadamente em cada dia; apenas a partir da 2ª ocorrência."
    ws.merge_cells("A48:N48")
    ws["A48"].fill = PatternFill("solid", fgColor=azul_claro)
    ws["A48"].font = Font(italic=True, color=azul)

    borda = Border(bottom=Side(style="thin", color="CBD5E1"))
    for linha_celulas in ws.iter_rows(min_row=11, max_row=15, min_col=1, max_col=3):
        for celula in linha_celulas:
            celula.border = borda
    for coluna, largura in {"A": 25, "B": 18, "C": 14, "D": 20, "E": 14, "F": 16, "G": 20, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14}.items():
        ws.column_dimensions[coluna].width = largura
    ws.freeze_panes = "A4"
    ws.print_area = "A1:N49"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def gerar_excel(registros: list[RegistroValidado], saida: Path, momento: datetime) -> pd.DataFrame:
    df = pd.DataFrame([registro.to_dict() for registro in registros])
    saida.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Resumo", index=False)
        df.to_excel(writer, sheet_name="Todos", index=False)
        for classificacao, aba in ABAS.items():
            df[df["Classificação"] == classificacao].to_excel(writer, sheet_name=aba, index=False)

    wb = load_workbook(saida)
    montar_resumo(wb["Resumo"], df, momento)
    for numero, aba in enumerate(("Todos", *ABAS.values()), start=1):
        estilizar_tabela(wb[aba], f"Tabela{numero}")
    wb.active = wb.sheetnames.index("Resumo")
    wb.calculation.fullCalcOnLoad = True
    wb.save(saida)
    return df


def gerar_pdf_resumo(df: pd.DataFrame, saida: Path) -> bool:
    """Cria uma réplica estática do painel para impressão/entrega."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return False
    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    diario = df.groupby(["Data de Referência", "Classificação"]).size().unstack(fill_value=0).reindex(columns=CLASSIFICACOES, fill_value=0)
    diario.index = pd.to_datetime(diario.index, format="%d/%m/%Y")
    diario = diario.sort_index()
    fig = plt.figure(figsize=(16, 9), facecolor="#F7FAFC")
    grade = fig.add_gridspec(3, 5, height_ratios=[0.7, 2.7, 0.8], hspace=0.45, wspace=0.5)
    fig.suptitle("CONFERÊNCIA DE LOTES · PAINEL EXECUTIVO", fontsize=22, fontweight="bold", color="#17365D", y=0.97)
    cards = [("TOTAL", len(df), "#17365D")] + [(n.upper(), int(contagens[n]), "#" + CORES[n]) for n in CLASSIFICACOES]
    for i, (nome, valor, cor) in enumerate(cards):
        ax = fig.add_subplot(grade[0, i]); ax.axis("off")
        pct = "" if nome == "TOTAL" else f"\n{valor / len(df):.1%}"
        ax.text(0.5, 0.5, f"{nome}\n{valor}{pct}", ha="center", va="center", color="white", fontsize=12, fontweight="bold", bbox=dict(boxstyle="round,pad=0.8", facecolor=cor, edgecolor=cor))
    ax1 = fig.add_subplot(grade[1, :2])
    ax1.pie(contagens, labels=CLASSIFICACOES, autopct="%1.1f%%", startangle=90, colors=["#" + CORES[n] for n in CLASSIFICACOES], wedgeprops=dict(width=0.42, edgecolor="white"))
    ax1.set_title("Distribuição dos registros", fontweight="bold", color="#17365D")
    ax2 = fig.add_subplot(grade[1, 2:])
    eixo_x = diario.index.strftime("%d/%m")
    ax2.plot(eixo_x, diario["Divergência"], marker="o", label="Divergências", color="#F59E0B")
    ax2.plot(eixo_x, diario["Ambíguo"], marker="o", label="Ambíguos", color="#8B5CF6")
    total_problemas = diario[["Divergência", "Ambíguo", "Erro de Entrada"]].sum(axis=1)
    ax2.plot(eixo_x, total_problemas, marker="o", linewidth=2.5, label="Total de problemas", color="#EF4444")
    ax2.set_title("Evolução dos registros", fontweight="bold", color="#17365D")
    ax2.set_ylabel("Quantidade"); ax2.grid(axis="y", alpha=.25); ax2.legend()
    ax3 = fig.add_subplot(grade[2, :]); ax3.axis("off")
    ax3.text(0, .75, f"Corrigir na origem: {contagens['Erro de Entrada']}  |  Conciliar: {contagens['Divergência']}  |  Decisão humana: {contagens['Ambíguo']}", fontsize=13, fontweight="bold", color="#17365D")
    ax3.text(0, .25, "Duplicidades são avaliadas separadamente em cada dia e somente a partir da 2ª ocorrência.", fontsize=10, color="#475569")
    fig.savefig(saida, bbox_inches="tight")
    plt.close(fig)
    return True


def salvar_log(df: pd.DataFrame, caminho: Path, origem: Path, momento: datetime) -> None:
    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    problemas = int(contagens[["Divergência", "Ambíguo", "Erro de Entrada"]].sum())
    resultado_obtido = {nome: int(contagens[nome]) for nome in CLASSIFICACOES}
    aceite = len(df) == 250 and resultado_obtido == TOTAIS_GABARITO
    linhas = [
        "LOG DE EXECUÇÃO — CONFERÊNCIA DE LOTES",
        f"Data/hora: {momento.strftime('%d/%m/%Y %H:%M:%S')}",
        f"Arquivo de origem: {origem.resolve()}",
        f"Total processado: {len(df)}",
        *(f"{nome}: {int(contagens[nome])}" for nome in CLASSIFICACOES),
        f"Total de registros problemáticos (todas as categorias): {problemas}",
        "Gabarito: 150 válidos + 50 divergências + 20 ambíguos + 30 erros de entrada = 250.",
        "RN11: contagem reiniciada em cada aba diária; divergência somente a partir da 2ª ocorrência.",
        "Validação de aceite: " + ("APROVADA" if aceite else "REVISAR"),
    ]
    # BOM facilita a abertura correta em Bloco de Notas/PowerShell legados.
    caminho.write_text("\n".join(linhas) + "\n", encoding="utf-8-sig")


def localizar_entrada(argumento: str | None) -> Path:
    candidatos = [
        Path(argumento) if argumento else None,
        Path("data/input/inspecao_lotes_10dias.xlsx"),
        Path.home() / "Downloads" / "inspecao_lotes_10dias.xlsx",
    ]
    for candidato in candidatos:
        if candidato and candidato.exists():
            return candidato
    raise FileNotFoundError("Informe o caminho de inspecao_lotes_10dias.xlsx.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("entrada", nargs="?", help="Planilha de inspeções")
    parser.add_argument("--saida", default="reports/relatorio_conferencia_lotes.xlsx")
    args = parser.parse_args()
    origem, saida = localizar_entrada(args.entrada), Path(args.saida)
    momento = datetime.now()
    registros = ler_e_validar(origem)
    df = gerar_excel(registros, saida, momento)
    salvar_log(df, saida.with_name("log_execucao.txt"), origem, momento)
    pdf_ok = gerar_pdf_resumo(df, saida.with_name("dashboard_resumo.pdf"))
    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    print(f"Relatório: {saida.resolve()}")
    print(f"Total: {len(df)} | " + " | ".join(f"{n}: {int(contagens[n])}" for n in CLASSIFICACOES))
    print("PDF: gerado" if pdf_ok else "PDF: matplotlib não instalado; Excel gerado normalmente")


if __name__ == "__main__":
    main()
